import openpyxl


workbook_name = "WEEK 9 HFR REPORTs.xlsx"
data_sheet_name = "DATA"
sheets_to_check = ["CUMMULATIVES","HTS", "PMTCT-EID_HEI", "INDEX", "TB", "TB_PREV", "APPOINTMENT"]

def analyze_workbook():
 
  wb = openpyxl.load_workbook(workbook_name)

 
  data_sheet = wb[data_sheet_name]
  for row in range(4, 58):  
    b_value, c_value, d_value = data_sheet[row][1].value, data_sheet[row][2].value, data_sheet[row][3].value
    if not all([b_value, c_value, d_value]):  
      for col in range(3, 5): 
        data_sheet.cell(row, col).fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")

   
    if c_value and c_value > b_value:
      for col in range(1, 3):  
        data_sheet.cell(row, col).fill = openpyxl.styles.PatternFill("solid", fgColor="FFD700")
    if d_value and d_value > c_value:
      for col in range(2, 4):  
        data_sheet.cell(row, col).fill = openpyxl.styles.PatternFill("solid", fgColor="FFD700")
    elif d_value and d_value > b_value:
      for col in range(1, 4): 
        data_sheet.cell(row, col).fill = openpyxl.styles.PatternFill("solid", fgColor="FFD700")

  
  for sheet_name in sheets_to_check:
    sheet = wb[sheet_name]
    for row in range(3, 58): 
      av_value, aw_value, ax_value = sheet[row][21].value, sheet[row][22].value, sheet[row][23].value

     
      if not any([av_value, aw_value, ax_value]):
        for col in range(21, 24):
          sheet.cell(row, col).fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")

    
      if aw_value and aw_value < av_value:
        for col in range(21, 23):  
          sheet.cell(row, col).fill = openpyxl.styles.PatternFill("solid", fgColor="FFD700")
      if aw_value and (aw_value != ax_value):
        for col in range(22, 24):  
          sheet.cell(row, col).fill = openpyxl.styles.PatternFill("solid", fgColor="FFD700")

  wb.save(workbook_name)  
